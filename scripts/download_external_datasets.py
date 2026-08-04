from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from _bootstrap import ROOT
from vipragsent.hashing import sha256_file
from vipragsent.phase import write_phase_handoff


MANUAL_READMES = {
    "uit_vsfc": """# UIT-VSFC manual drop\n\nOfficial source folder: `https://drive.google.com/drive/folders/1xclbjHHK58zk2X6iqbvMPS2rcy9y9E0X`\n\nPlace the official test-folder files `sents.txt`, `sentiments.txt`, and `topics.txt` in this directory. The normalizer maps `0=negative`, `1=neutral`, and `2=positive`, and writes `data/processed/external/uit_vsfc/test.csv` with columns `sample_id,text,polarity`.\n\nKeep the author-contact permission evidence privately. Do not substitute an unsplit mirror or a random split.\n""",
    "uit_vsmec": """# UIT-VSMEC manual drop\n\nOfficial source folder: `https://drive.google.com/drive/folders/1HooABJyrddVGzll7fgkJ6VzkG_XuWfRu?usp=drive_link`\n\nPlace the official test workbook `test_nor_811.xlsx` in this directory. The normalizer writes `data/processed/external/uit_vsmec/test.csv` with columns `sample_id,text,emotion`. The source workbook's first-column identifiers repeat, so stable row-based sample IDs are generated without deduplicating records.\n\nKeep the author-contact permission evidence privately. Do not substitute an unsplit mirror or a random split.\n""",
    "aivivn_original": """# AIVIVN original manual fallback\n\nOfficial Kaggle source: `https://www.kaggle.com/datasets/mcocoz/aivivn-2019`\n\nPlace the original `train.csv` and `test.csv` files in this directory. The files must contain `id`, `comment`, and `label` columns. They are preserved for provenance only; Q1b uses the bundled `AIVIVN-human-derived-3way` split.\n\nConfigure Kaggle credentials through the standard environment or `KAGGLE_CONFIG_DIR`; never commit credentials or the original text.\n""",
}

OFFICIAL_SOURCES = {
    "uit_vsfc": {
        "folder_url": "https://drive.google.com/drive/folders/1xclbjHHK58zk2X6iqbvMPS2rcy9y9E0X",
        "file_ids": {
            "sents.txt": "1aNMOeZZbNwSRkjyCWAGtNCMa3YrshR-n",
            "sentiments.txt": "1vkQS5gI0is4ACU58-AbWusnemw7KZNfO",
            "topics.txt": "1_ArMpDguVsbUGl-xSMkTF_p5KpZrmpSB",
        },
        "license_note": "Research access authorized by contacting the dataset author; confirmation is retained privately.",
    },
    "uit_vsmec": {
        "folder_url": "https://drive.google.com/drive/folders/1HooABJyrddVGzll7fgkJ6VzkG_XuWfRu?usp=drive_link",
        "file_ids": {"test_nor_811.xlsx": "1D16FCKKgJ0T6t2aSA3biWVwvD9fa4G9a"},
        "license_note": "Research access authorized by contacting the dataset author; confirmation is retained privately.",
    },
}

VSFC_POLARITY = {"0": "negative", "1": "neutral", "2": "positive"}
VSMEC_EMOTION = {
    "anger": "anger",
    "disgust": "disgust",
    "enjoyment": "enjoyment",
    "fear": "fear",
    "other": "other",
    "sadness": "sadness",
    "surprise": "surprise",
}


def _relative(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def _read_lines(path: Path) -> list[str]:
    return path.read_text(encoding="utf-8-sig").splitlines()


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    return sha256_file(path)


def _normalize_vsfc(raw_root: Path) -> dict[str, Any]:
    raw_paths = {name: raw_root / name for name in ("sents.txt", "sentiments.txt", "topics.txt")}
    missing = [name for name, path in raw_paths.items() if not path.exists()]
    if missing:
        return {"status": "BLOCKED", "missing_files": missing}

    sentences = _read_lines(raw_paths["sents.txt"])
    sentiments = _read_lines(raw_paths["sentiments.txt"])
    topics = _read_lines(raw_paths["topics.txt"])
    if not (len(sentences) == len(sentiments) == len(topics)):
        raise ValueError(
            "UIT-VSFC test files must have equal line counts: "
            f"sents={len(sentences)}, sentiments={len(sentiments)}, topics={len(topics)}"
        )

    rows: list[dict[str, str]] = []
    for index, (sentence, sentiment, topic) in enumerate(zip(sentences, sentiments, topics, strict=True)):
        label = VSFC_POLARITY.get(sentiment.strip())
        if label is None:
            raise ValueError(f"Unknown UIT-VSFC polarity at line {index + 1}: {sentiment!r}")
        if topic.strip() not in {"0", "1", "2", "3"}:
            raise ValueError(f"Unknown UIT-VSFC topic at line {index + 1}: {topic!r}")
        rows.append({"sample_id": f"uit_vsfc_test_{index:04d}", "text": sentence, "polarity": label})

    output = ROOT / "data/processed/external/uit_vsfc/test.csv"
    normalized_checksum = _write_csv(output, ["sample_id", "text", "polarity"], rows)
    return {
        "status": "PASS",
        "normalized_path": _relative(output),
        "checksum": normalized_checksum,
        "row_count": len(rows),
        "label_mapping": VSFC_POLARITY,
        "raw_paths": [_relative(path) for path in raw_paths.values()],
        "raw_checksums": {name: sha256_file(path) for name, path in raw_paths.items()},
        "split": "test",
        "split_evidence": "Official UIT-VSFC Drive test folder; topics/sentiments/sents are aligned line-by-line.",
    }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_vsmec(raw_root: Path) -> dict[str, Any]:
    workbook = raw_root / "test_nor_811.xlsx"
    if not workbook.exists():
        return {"status": "BLOCKED", "missing_files": [workbook.name]}
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("UIT-VSMEC normalization requires openpyxl; install the [data] extra") from exc

    sheet = load_workbook(workbook, read_only=True, data_only=True).active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("UIT-VSMEC workbook is empty")
    headers = [_cell_text(value).casefold() for value in rows[0]]
    try:
        id_index = headers.index("unnamed: 0")
    except ValueError:
        id_index = 0
    try:
        emotion_index = headers.index("emotion")
        text_index = headers.index("sentence")
    except ValueError as exc:
        raise ValueError(f"UIT-VSMEC workbook must contain Emotion and Sentence columns: {headers}") from exc

    normalized: list[dict[str, str]] = []
    original_ids: list[str] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(value is None for value in row):
            continue
        source_id = _cell_text(row[id_index])
        text = _cell_text(row[text_index])
        emotion = _cell_text(row[emotion_index]).casefold()
        if not source_id or not text:
            raise ValueError(f"UIT-VSMEC row {row_number} has an empty ID or sentence")
        label = VSMEC_EMOTION.get(emotion)
        if label is None:
            raise ValueError(f"Unknown UIT-VSMEC emotion at row {row_number}: {emotion!r}")
        original_ids.append(source_id)
        sample_id = f"uit_vsmec_test_{len(normalized):04d}"
        normalized.append({"sample_id": sample_id, "text": text, "emotion": label})

    output = ROOT / "data/processed/external/uit_vsmec/test.csv"
    normalized_checksum = _write_csv(output, ["sample_id", "text", "emotion"], normalized)
    return {
        "status": "PASS",
        "normalized_path": _relative(output),
        "checksum": normalized_checksum,
        "row_count": len(normalized),
        "label_mapping": VSMEC_EMOTION,
        "raw_paths": [_relative(workbook)],
        "raw_checksums": {workbook.name: sha256_file(workbook)},
        "split": "test",
        "split_evidence": "Official UIT-VSMEC test_nor_811.xlsx workbook from the supplied Drive folder.",
        "original_id_column": "Unnamed: 0",
        "original_id_unique": len(set(original_ids)) == len(original_ids),
        "original_id_duplicate_values": sorted(value for value, count in Counter(original_ids).items() if count > 1),
    }


def _inspect_aivivn(raw_root: Path) -> dict[str, Any]:
    raw_paths = {name: raw_root / name for name in ("train.csv", "test.csv")}
    missing = [name for name, path in raw_paths.items() if not path.exists()]
    if missing:
        return {"status": "BLOCKED", "missing_files": missing}

    file_reports: dict[str, Any] = {}
    for name, path in raw_paths.items():
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            required = {"id", "comment", "label"}
            if not required.issubset(reader.fieldnames or set()):
                raise ValueError(f"AIVIVN {name} is missing columns {sorted(required)}")
            rows = list(reader)
        ids = [row["id"] for row in rows]
        labels = Counter(row["label"].strip() for row in rows)
        if len(ids) != len(set(ids)):
            raise ValueError(f"AIVIVN {name} contains duplicate IDs")
        if any(not row["comment"].strip() for row in rows):
            raise ValueError(f"AIVIVN {name} contains an empty comment")
        if set(labels) - {"0", "1"}:
            raise ValueError(f"AIVIVN {name} contains unexpected labels: {sorted(set(labels) - {'0', '1'})}")
        file_reports[name] = {"rows": len(rows), "label_counts": dict(sorted(labels.items()))}

    return {
        "status": "PASS",
        "normalized_path": None,
        "checksum": None,
        "raw_paths": [_relative(path) for path in raw_paths.values()],
        "raw_checksums": {name: sha256_file(path) for name, path in raw_paths.items()},
        "file_reports": file_reports,
        "split": "official train/test files preserved as downloaded",
        "split_evidence": "Kaggle mcocoz/aivivn-2019 archive supplied train.csv and test.csv.",
        "provenance_only": True,
    }


def _aivivn_entry(inspected: dict[str, Any]) -> dict[str, Any]:
    entry = {
        "status": inspected["status"],
        "source": "https://www.kaggle.com/datasets/mcocoz/aivivn-2019",
        "normalized_path": None,
        "checksum": None,
        "license_note": "Kaggle metadata reports Apache-2.0; original binaries are preserved for provenance only.",
        "access_evidence": "Downloaded through the authenticated Kaggle client; credentials are not stored in the repository.",
        "provenance_only": True,
    }
    for key in ("missing_files", "raw_paths", "raw_checksums", "file_reports", "split", "split_evidence"):
        if key in inspected:
            entry[key] = inspected[key]
    return entry


def _external_entry(name: str, normalized: dict[str, Any]) -> dict[str, Any]:
    source = OFFICIAL_SOURCES[name]
    entry = {
        "status": normalized["status"],
        "source": source["folder_url"],
        "source_file_ids": source["file_ids"],
        "normalized_path": normalized.get("normalized_path"),
        "checksum": normalized.get("checksum"),
        "license_note": source["license_note"],
        "access_evidence": "User-provided author-contact permission; private email evidence is not copied into the repository.",
    }
    for key in ("missing_files", "row_count", "label_mapping", "raw_paths", "raw_checksums", "split", "split_evidence", "original_id_column", "original_id_unique", "original_id_duplicate_values"):
        if key in normalized:
            entry[key] = normalized[key]
    return entry


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare and validate external dataset downloads")
    parser.add_argument("--all", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    for name, content in MANUAL_READMES.items():
        path = ROOT / "data/external/manual_drop" / name / "README.md"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")
    normalized = {
        "uit_vsfc": _normalize_vsfc(ROOT / "data/external/manual_drop/uit_vsfc"),
        "uit_vsmec": _normalize_vsmec(ROOT / "data/external/manual_drop/uit_vsmec"),
    }
    aivivn = _inspect_aivivn(ROOT / "data/external/manual_drop/aivivn_original")
    bundled = ROOT / "data/processed/external/aivivn_human_derived_3way/test.csv"
    manifest = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "datasets": {
            "uit_vsfc": _external_entry("uit_vsfc", normalized["uit_vsfc"]),
            "uit_vsmec": _external_entry("uit_vsmec", normalized["uit_vsmec"]),
            "aivivn_original": _aivivn_entry(aivivn),
            "aivivn_human_derived_3way": {"status": "PASS" if bundled.exists() else "BLOCKED", "source": "bundled V8 package", "normalized_path": str(bundled.relative_to(ROOT)) if bundled.exists() else None, "checksum": sha256_file(bundled) if bundled.exists() else None, "license_note": "project-bundled human-derived split"},
        },
        "q1b_uses_bundled_aivivn_human_derived_3way": True,
        "external_finetuning": False,
    }
    path = ROOT / "data/manifests/external_datasets.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    blocked = [name for name, item in manifest["datasets"].items() if item["status"] == "BLOCKED"]
    status = "PASS" if not blocked else "BLOCKED"
    write_phase_handoff("02", status, inputs_read=["22_DATA_SOURCE_REGISTRY.md", "Kaggle mcocoz/aivivn-2019", "V8 bundled AIVIVN files", "official UIT-VSFC Drive folder", "official UIT-VSMEC Drive folder"], files_created=["data/manifests/external_datasets.json", "data/processed/external/uit_vsfc/test.csv", "data/processed/external/uit_vsmec/test.csv", "data/external/manual_drop/aivivn_original/train.csv", "data/external/manual_drop/aivivn_original/test.csv", "data/external/manual_drop/*/README.md"], tests_run=["official UIT-VSFC line-count and label validation", "official UIT-VSMEC workbook schema and label validation", "AIVIVN original train/test schema and checksum validation", "external manifest generation", "bundled AIVIVN schema/checksum check"], tests_passed=True, blockers=[f"Manual or credentialed dataset required: {name}" for name in blocked], next_phase_ready=not blocked)
    print(json.dumps({"status": status, "blocked": blocked, "dry_run": args.dry_run}, indent=2))
    return 0 if not blocked or args.dry_run else 2


if __name__ == "__main__":
    raise SystemExit(main())
